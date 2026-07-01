"""Publish all CSV tables to one Excel workbook and one Google Sheets file."""

from __future__ import annotations

from pathlib import Path
import re
from typing import Any

import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def publish_table_workbook(table_directory: Path, output_path: Path) -> Path:
    """Write every CSV table to a single formatted Excel workbook."""
    table_paths = _select_table_paths(table_directory)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _metadata_frame().to_excel(writer, sheet_name="Metadata", index=False)
        index = pd.DataFrame(
            [
                {"sheet": _excel_sheet_name(path, number), "source_file": path.name, "rows": len(_read_table(path))}
                for number, path in enumerate(table_paths, 1)
            ]
        )
        index.to_excel(writer, sheet_name="Index", index=False)
        for number, path in enumerate(table_paths, 1):
            frame = _read_table(path)
            sheet_name = _excel_sheet_name(path, number)
            metadata = _table_metadata_rows(path)
            pd.DataFrame(metadata).to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(metadata) + 1)

        for worksheet in writer.book.worksheets:
            _format_excel_sheet(worksheet)
    return output_path


def publish_google_sheets_tables(
    sheets_service: Any,
    settings_path: Path,
    settings: dict[str, Any],
    table_directory: Path,
) -> str:
    """Create or update one Google Sheets workbook containing all CSV tables."""
    spreadsheet_id = _ensure_tables_spreadsheet(sheets_service, settings_path, settings)
    table_paths = _select_table_paths(table_directory)
    metadata = _execute_with_retry(
        sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id)
    )
    existing = metadata.get("sheets", [])
    requests = []
    if existing:
        first_id = existing[0]["properties"]["sheetId"]
        requests.append({"updateSheetProperties": {"properties": {"sheetId": first_id, "title": "Metadata"}, "fields": "title"}})
        for sheet in existing[1:]:
            requests.append({"deleteSheet": {"sheetId": sheet["properties"]["sheetId"]}})
    requests.append({"addSheet": {"properties": {"title": "Index"}}})
    for number, path in enumerate(table_paths, 1):
        requests.append({"addSheet": {"properties": {"title": _sheets_tab_name(path, number)}}})
    if requests:
        sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()

    index_values = [["Sheet", "Source file", "Rows"]]
    updates = [
        {
            "range": "'Metadata'!A1",
            "majorDimension": "ROWS",
            "values": _frame_values(_metadata_frame()),
        }
    ]
    for number, path in enumerate(table_paths, 1):
        frame = _read_table(path)
        sheet_name = _sheets_tab_name(path, number)
        index_values.append([sheet_name, path.name, len(frame)])
        metadata_rows = _table_metadata_rows(path)
        updates.append(
            {
                "range": f"'{sheet_name}'!A1",
                "majorDimension": "ROWS",
                "values": metadata_rows + [[]] + _frame_values(frame),
            }
        )
    updates.insert(1, {"range": "'Index'!A1", "majorDimension": "ROWS", "values": index_values})
    sheets_service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "USER_ENTERED", "data": updates},
    ).execute()
    _format_google_sheets(sheets_service, spreadsheet_id)
    return spreadsheet_id


def _ensure_tables_spreadsheet(sheets_service: Any, settings_path: Path, settings: dict[str, Any]) -> str:
    spreadsheet_id = str(settings["google"].get("tables_spreadsheet_id", "")).strip()
    if spreadsheet_id:
        return spreadsheet_id
    title = str(settings["google"].get("tables_spreadsheet_title") or "MapBiomas CTrees Tables")
    created = sheets_service.spreadsheets().create(body={"properties": {"title": title}}).execute()
    spreadsheet_id = str(created["spreadsheetId"])
    settings["google"]["tables_spreadsheet_id"] = spreadsheet_id
    data = yaml.safe_load(settings_path.read_text(encoding="utf-8"))
    data.setdefault("google", {})["tables_spreadsheet_id"] = spreadsheet_id
    settings_path.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=False), encoding="utf-8")
    return spreadsheet_id


def _select_table_paths(table_directory: Path) -> list[Path]:
    priority = [
        "agreement_metrics.csv",
        "fcbm_comparison_metrics.csv",
        "change_area_forest_to_nonforest.csv",
        "ChangeArea_ForestToNonForest_30m.csv",
    ]
    priority_paths = [table_directory / name for name in priority if (table_directory / name).exists()]
    remaining = [path for path in sorted(table_directory.glob("*.csv")) if path not in set(priority_paths)]
    return priority_paths + remaining


def _metadata_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "scheme": "VT0007 four-class interpretation",
                "classes": "1 Stable non-forest; 2 Stable forest; 3 Deforested first half of HRP; 4 Deforested second half of HRP",
                "fcbm_indices": "1-4 -> 1; 5 -> 2; 6-7 -> 3; 8 -> 4",
                "source": "VMD0055 v1.1, Table 15; VT0007 v1.0",
            },
            {
                "scheme": "UDef-A risk map binary interpretation",
                "classes": "Forest/non-forest at T1, T2, and T3 with period-specific FCBM groupings",
                "fcbm_indices": "T1 forest 5-8; T2 forest 5 and 8; T3 forest 5",
                "source": "VT0007 v1.0, Table 1 and Section Data Requirements",
            },
            {
                "scheme": "Accuracy assessment three-class interpretation",
                "classes": "1 Non-forest at HRP end; 2 Forest at HRP end; 3 Deforested within HRP",
                "fcbm_indices": "1,3 -> 1; 2,4,5 -> 2; 6-8 -> 3",
                "source": "VMD0055 v1.1, Table 16",
            },
        ]
    )


def _table_metadata_rows(path: Path) -> list[list[str]]:
    return [
        ["Product", path.stem],
        ["Source file", path.name],
        ["Class scheme", _class_scheme_for_table(path)],
        ["Verra reference", _verra_reference_for_table(path)],
    ]


def _class_scheme_for_table(path: Path) -> str:
    name = path.stem.lower()
    if "fcbm_comparison_metrics" in name:
        return "Accuracy metrics for CTrees FCBM versus MapBiomas-derived FCBM comparisons"
    if "fcbm_comparison" in name and "fcbm_xtab" in name:
        return "CTrees FCBM index cross-tabulated against MapBiomas-derived FCBM, with metrics evaluated under the accuracy assessment three-class scheme"
    if "riskmap_xtab" in name:
        return "CTrees UDef-A risk-map binary forest/non-forest product cross-tabulated against MapBiomas binary forest/non-forest for the corresponding year"
    if "fcbm_comparison" in name or "derivedbinary_xtab" in name:
        return "Paired CTrees-derived and MapBiomas-derived binary UDef-A product agreement"
    if "changearea" in name or "foresttononforest" in name or "change_area" in name:
        return "Forest-to-nonforest change area, deforestation classes aggregated where required"
    if "xtab" in name or "crosstab" in name:
        return "Configured CTrees reference class scheme cross-tabulated against MapBiomas persistence classes"
    if "agreement" in name:
        return "Accuracy metrics derived from configured forest, non-forest, and change groups"
    if "reclassification" in name:
        return "MapBiomas forest/non-forest/excluded analytical reclassification"
    return "Analytical table generated from MapBiomas and CTrees workflow outputs"


def _verra_reference_for_table(path: Path) -> str:
    name = path.stem.lower()
    if "fcbm_comparison_metrics" in name:
        return "VMD0055 v1.1, Table 16; VT0007 v1.0, Table 1"
    if "fcbm_comparison" in name and "fcbm_xtab" in name:
        return "VMD0055 v1.1, Table 16 for accuracy metrics; VT0007 v1.0, Table 1 for the eight-index FCBM"
    if "riskmap_xtab" in name:
        return "VT0007 v1.0, Table 1 and Section Data Requirements for Test, HRP, and Validity risk-map binary forest definitions"
    if "fcbm_comparison" in name or "derivedbinary_xtab" in name:
        return "VT0007 v1.0, Table 1 and Section Data Requirements for UDef-A binary FCBM products"
    if "changearea" in name or "foresttononforest" in name or "change_area" in name:
        return "VT0007 v1.0, Section Data Requirements; VMD0055 v1.1, Table 15"
    if "xtab" in name or "crosstab" in name:
        return "VMD0055 v1.1, Table 15 or Table 16, according to the table-specific CTrees interpretation"
    if "agreement" in name:
        return "VMD0055 v1.1, Table 16 for accuracy assessment metrics"
    return "VMD0055 v1.1 and VT0007 v1.0 analytical traceability"


def _frame_values(frame: pd.DataFrame) -> list[list[Any]]:
    clean = frame.fillna("")
    return [list(clean.columns), *clean.astype(object).values.tolist()]


def _read_table(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path)
    except pd.errors.EmptyDataError:
        return pd.DataFrame([{"message": "No rows were available for this optional table at publication time."}])


def _excel_sheet_name(path: Path, number: int) -> str:
    base = _clean_sheet_name(path.stem)
    return f"{number:02d}_{base}"[:31]


def _sheets_tab_name(path: Path, number: int) -> str:
    base = _clean_sheet_name(path.stem)
    return f"{number:03d}_{base}"[:100]


def _clean_sheet_name(value: str) -> str:
    text = re.sub(r"[\[\]\*\?/\\:]", "_", value)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_") or "table"


def _format_excel_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="F1F3F4")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="202124")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _format_google_sheets(sheets_service: Any, spreadsheet_id: str) -> None:
    metadata = sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    requests = []
    for sheet in metadata.get("sheets", []):
        sheet_id = sheet["properties"]["sheetId"]
        requests.extend(
            [
                {"updateSheetProperties": {"properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}}, "fields": "gridProperties.frozenRowCount"}},
                {
                    "repeatCell": {
                        "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.945, "green": 0.953, "blue": 0.957},
                                "textFormat": {"bold": True},
                                "horizontalAlignment": "CENTER",
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
                    }
                },
                {"autoResizeDimensions": {"dimensions": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 20}}},
            ]
        )
    if requests:
        sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()
import time

from googleapiclient.errors import HttpError


def _execute_with_retry(request, attempts: int = 5):
    """Execute a Google API request with retry for transient service errors."""
    last_error = None
    for attempt in range(attempts):
        try:
            return request.execute()
        except HttpError as exc:
            last_error = exc
            status = getattr(exc.resp, "status", None)
            if status not in {429, 500, 502, 503, 504} or attempt == attempts - 1:
                raise
            time.sleep(min(2 ** attempt, 30))
    raise last_error


ESSENTIAL_TABLE_PREFIXES = (
    "area_",
    "Area_",
    "change_area",
    "ChangeArea_",
    "forest_change",
    "ForestChange_",
    "temporal_consistency",
    "municipal_",
    "MapBiomas_",
)

ESSENTIAL_TABLE_NAMES = {
    "agreement_metrics.csv",
    "mapbiomas_reclassification_schema.csv",
    "change_area_by_interval.csv",
    "change_area_forest_to_nonforest.csv",
    "forest_change_area_timeseries.csv",
    "temporal_consistency_reversals.csv",
}


def _is_essential_table(path: Path) -> bool:
    """Return True for tables used in final reporting or interpretation."""
    return path.name in ESSENTIAL_TABLE_NAMES or path.name.startswith(ESSENTIAL_TABLE_PREFIXES)
